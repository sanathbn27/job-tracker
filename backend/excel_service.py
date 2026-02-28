import os
import openpyxl
from datetime import datetime
from dotenv import load_dotenv

load_dotenv()

EXCEL_FILE = 'job_tracker.xlsx'

HEADERS = [
    'ID', 'Company', 'Role', 'Location',
    'Date Applied', 'Date Responded', 'Days Taken',
    'Status', 'Interview Round', 'Source',
    'Email Thread ID', 'Notes'
]


def save_workbook(wb) -> bool:
    """
    Saves workbook with error handling for when file is open in Excel.
    Returns True if saved successfully, False if file is locked.
    """
    try:
        wb.save(EXCEL_FILE)
        return True
    except PermissionError:
        print(f"Warning: Could not save Excel file — file is currently open in Excel.")
        print(f"Please close {EXCEL_FILE} and the data will be saved on next email.")
        return False
    except Exception as e:
        print(f"Excel save error: {e}")
        return False

def get_or_create_workbook():
    """
    Opens existing Excel file or creates a new one with headers.
    """
    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Job Tracker'
        # Write headers
        for col, header in enumerate(HEADERS, 1):
            ws.cell(row=1, column=col, value=header)
        save_workbook(wb)
        print(f"Created new Excel file: {EXCEL_FILE}")

    return wb, ws


def find_matching_row_excel(ws, company: str, role: str) -> int | None:
    """
    Finds matching row in Excel by company and role.
    Returns row number (1-based) or None.
    """
    company_lower = company.lower().strip()
    role_lower = role.lower().strip()

    for row in ws.iter_rows(min_row=2, values_only=False):
        row_company = str(row[1].value or '').lower().strip()  # Column B
        row_role = str(row[2].value or '').lower().strip()     # Column C
        row_status = str(row[7].value or '')                   # Column H

        company_match = (
            company_lower in row_company or
            row_company in company_lower
        )
        role_match = (
            role_lower in row_role or
            row_role in role_lower
        )

        if company_match and role_match and row_status == 'Applied':
            return row[0].row

    return None


def create_new_row_excel(parsed_data: dict, note_override: str = '') -> bool:
    """Creates a new row in Excel file."""
    try:
        wb, ws = get_or_create_workbook()

        # Get next job ID from last row
        last_row = ws.max_row
        if last_row == 1:
            job_num = 1
        else:
            last_id = ws.cell(row=last_row, column=1).value or 'JOB000'
            try:
                job_num = int(str(last_id)[3:]) + 1
            except ValueError:
                job_num = last_row

        job_id = f'JOB{str(job_num).zfill(3)}'
        notes = note_override if note_override else parsed_data.get('notes', '')

        # Calculate days taken
        from backend.sheets_service import calculate_days_taken
        days_taken = calculate_days_taken(
            parsed_data.get('date_applied', ''),
            parsed_data.get('date_responded', '')
        )

        new_row = [
            job_id,
            parsed_data.get('company', ''),
            parsed_data.get('role', ''),
            parsed_data.get('location', ''),
            parsed_data.get('date_applied', ''),
            parsed_data.get('date_responded', ''),
            days_taken,
            parsed_data.get('status', ''),
            parsed_data.get('interview_round', ''),
            parsed_data.get('source', ''),
            parsed_data.get('thread_id', ''),
            notes
        ]

        ws.append(new_row)
        save_workbook(wb)
        print(f"Excel: Created row {job_id} - {parsed_data.get('company')}")
        return True

    except Exception as e:
        print(f"Excel create error: {e}")
        return False


def update_existing_row_excel(parsed_data: dict) -> bool:
    """Updates existing row in Excel file."""
    try:
        wb, ws = get_or_create_workbook()

        company = parsed_data.get('company', '')
        role = parsed_data.get('role', '')
        row_num = find_matching_row_excel(ws, company, role)

        if not row_num:
            # No match — create new row with note
            return create_new_row_excel(
                parsed_data,
                note_override=f"No confirmation email received. {parsed_data.get('notes', '')}"
            )

        # Get existing date applied for days calculation
        date_applied = ws.cell(row=row_num, column=5).value or ''

        from backend.sheets_service import calculate_days_taken
        days_taken = calculate_days_taken(
            str(date_applied),
            parsed_data.get('date_responded', '')
        )

        # Update cells
        ws.cell(row=row_num, column=6).value = parsed_data.get('date_responded', '')
        ws.cell(row=row_num, column=7).value = days_taken
        ws.cell(row=row_num, column=8).value = parsed_data.get('status', '')
        if parsed_data.get('interview_round'):
            ws.cell(row=row_num, column=9).value = parsed_data.get('interview_round')
        if parsed_data.get('notes'):
            ws.cell(row=row_num, column=12).value = parsed_data.get('notes')

        save_workbook(wb)
        print(f"Excel: Updated row {row_num} - {company} → {parsed_data.get('status')}")
        return True

    except Exception as e:
        print(f"Excel update error: {e}")
        return False


def process_parsed_email_excel(parsed_data: dict) -> bool:
    """
    Main function for Excel — mirrors sheets_service logic.
    """
    status = parsed_data.get('status', '')

    if status == 'Applied':
        return create_new_row_excel(parsed_data)
    elif status in ['Rejected', 'Interview', 'Offer']:
        return update_existing_row_excel(parsed_data)
    else:
        return False